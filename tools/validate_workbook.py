#!/usr/bin/env python3
"""Validate final MVA workbooks with public, reproducible dependencies."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ERROR_PATTERN = re.compile(r"#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A)", re.IGNORECASE)
LANE_PATTERN = re.compile(r"\blane\b", re.IGNORECASE)
REQUIRED_TREND_TITLES = {"Vulnerabilities Discovered", "Vulnerabilities Remediated"}


def chart_title(chart: object) -> str:
    """Extract an embedded chart title without relying on private XML parsing."""
    title = getattr(chart, "title", None)
    rich = getattr(getattr(title, "tx", None), "rich", None)
    paragraphs = getattr(rich, "p", []) or []
    return "".join(
        getattr(run, "t", "") or ""
        for paragraph in paragraphs
        for run in (getattr(paragraph, "r", []) or [])
    )


def validate(input_path: Path, output_dir: Path, required_sheets: list[str]) -> dict:
    formulas = load_workbook(input_path, data_only=False, read_only=False)
    cached = load_workbook(input_path, data_only=True, read_only=False)
    formula_errors: list[dict[str, str]] = []
    lane_matches: list[dict[str, str]] = []
    sheet_summary: list[dict[str, object]] = []
    missing_trend_charts: list[dict[str, object]] = []

    for sheet in formulas.worksheets:
        cached_sheet = cached[sheet.title]
        for row in sheet.iter_rows():
            for cell in row:
                text = "" if cell.value is None else str(cell.value)
                cached_value = cached_sheet[cell.coordinate].value
                cached_text = "" if cached_value is None else str(cached_value)
                if ERROR_PATTERN.search(text) or ERROR_PATTERN.search(cached_text):
                    formula_errors.append(
                        {
                            "address": f"{sheet.title}!{cell.coordinate}",
                            "formula": text,
                            "cachedValue": cached_text,
                        }
                    )
                if LANE_PATTERN.search(text):
                    lane_matches.append({"address": f"{sheet.title}!{cell.coordinate}", "value": text})

        chart_inventory = [
            {"type": type(chart).__name__, "title": chart_title(chart)} for chart in sheet._charts
        ]
        line_chart_titles = {
            chart["title"] for chart in chart_inventory if chart["type"] == "LineChart"
        }
        if sheet.title in {"Executive Dashboard", "Monthly Dashboard"}:
            missing_titles = sorted(REQUIRED_TREND_TITLES - line_chart_titles)
            if missing_titles:
                missing_trend_charts.append(
                    {"sheet": sheet.title, "missingTitles": missing_titles}
                )

        sheet_summary.append(
            {
                "name": sheet.title,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "mergedRanges": len(sheet.merged_cells.ranges),
                "frozenAt": str(sheet.freeze_panes or ""),
                "charts": chart_inventory,
            }
        )

    missing_sheets = [name for name in required_sheets if name not in formulas.sheetnames]
    checks = {
        "xlsx_reopens": bool(formulas.sheetnames),
        "required_sheets_present": not missing_sheets,
        "formula_errors_absent": not formula_errors,
        "obsolete_lane_absent": not lane_matches,
        "required_line_charts_present": not missing_trend_charts,
    }
    result = {
        "input": str(input_path),
        "outputDir": str(output_dir),
        "status": "PASS" if all(checks.values()) else "FAIL",
        "checks": checks,
        "sheets": sheet_summary,
        "missingSheets": missing_sheets,
        "formulaErrors": formula_errors,
        "obsoleteLaneMatches": lane_matches,
        "missingTrendCharts": missing_trend_charts,
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "validation.json").write_text(json.dumps(result, indent=2), encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate an MVA Excel workbook.")
    parser.add_argument("--input", default="output/excel/mva_crowdstrike_final_team_sample.xlsx")
    parser.add_argument("--output", default="output/validation/workbook")
    parser.add_argument(
        "--sheets",
        default="Executive Dashboard,Monthly Dashboard,Adhoc Dashboard",
        help="Comma-separated required worksheet names.",
    )
    args = parser.parse_args()
    required_sheets = [value.strip() for value in args.sheets.split(",") if value.strip()]
    result = validate(Path(args.input), Path(args.output), required_sheets)
    print(json.dumps(result, indent=2))
    return 0 if result["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
