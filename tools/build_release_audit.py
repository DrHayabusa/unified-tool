#!/usr/bin/env python3
"""Build a reproducible audit for the final MVA handover folder."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
FINAL = ROOT / "final"
REQUIRED_SHEETS = {"Executive Dashboard", "Monthly Dashboard", "Adhoc Dashboard"}
REQUIRED_TRENDS = {"Vulnerabilities Discovered", "Vulnerabilities Remediated"}
FORMULA_ERROR = re.compile(r"#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A)", re.IGNORECASE)
LANE = re.compile(r"\blane\b", re.IGNORECASE)
SECRET_PATTERNS = [
    re.compile(rb"nvapi-[A-Za-z0-9_-]{20,}"),
    re.compile(rb"sk-(?:or-v1-)?[A-Za-z0-9_-]{20,}"),
]


def chart_title(chart: object) -> str:
    title = getattr(chart, "title", None)
    rich = getattr(getattr(title, "tx", None), "rich", None)
    paragraphs = getattr(rich, "p", []) or []
    return "".join(
        getattr(run, "t", "") or ""
        for paragraph in paragraphs
        for run in (getattr(paragraph, "r", []) or [])
    )


def audit_workbook(path: Path) -> tuple[dict, list[str]]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    errors: list[str] = []
    lane_matches = 0
    formula_errors = 0
    chart_inventory: dict[str, list[dict[str, str]]] = {}

    for sheet in workbook.worksheets:
        inventory = [
            {"type": type(chart).__name__, "title": chart_title(chart)}
            for chart in sheet._charts
        ]
        chart_inventory[sheet.title] = inventory
        if sheet.title in {"Executive Dashboard", "Monthly Dashboard"}:
            line_titles = {item["title"] for item in inventory if item["type"] == "LineChart"}
            missing = sorted(REQUIRED_TRENDS - line_titles)
            if missing:
                errors.append(f"{path.name}: {sheet.title} missing line charts: {', '.join(missing)}")

        for row in sheet.iter_rows():
            for cell in row:
                text = "" if cell.value is None else str(cell.value)
                formula_errors += int(bool(FORMULA_ERROR.search(text)))
                lane_matches += int(bool(LANE.search(text)))

    missing_sheets = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing_sheets:
        errors.append(f"{path.name}: missing sheets: {', '.join(missing_sheets)}")
    if formula_errors:
        errors.append(f"{path.name}: {formula_errors} formula errors")
    if lane_matches:
        errors.append(f"{path.name}: {lane_matches} obsolete Lane matches")

    return {
        "file": path.name,
        "sheets": workbook.sheetnames,
        "lane_matches": lane_matches,
        "formula_errors": formula_errors,
        "charts": chart_inventory,
    }, errors


def audit_pdf(path: Path) -> dict:
    reader = PdfReader(path)
    links = 0
    for page in reader.pages:
        for annotation_ref in page.get("/Annots", []) or []:
            annotation = annotation_ref.get_object()
            links += int(annotation.get("/Subtype") == "/Link")
    return {"file": path.name, "pages": len(reader.pages), "clickable_links": links}


def audit_csv(path: Path) -> dict:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader, [])
        rows = sum(1 for _ in reader)
    return {"file": str(path.relative_to(FINAL)), "rows": rows, "columns": len(header)}


def find_secret_leaks() -> list[str]:
    leaks: list[str] = []
    for path in FINAL.rglob("*"):
        if not path.is_file():
            continue
        data = path.read_bytes()
        if any(pattern.search(data) for pattern in SECRET_PATTERNS):
            leaks.append(str(path.relative_to(FINAL)))
    return leaks


def write_checksums() -> None:
    destination = FINAL / "SHA256SUMS.txt"
    entries: list[str] = []
    for path in sorted(FINAL.rglob("*")):
        if not path.is_file() or path == destination:
            continue
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        entries.append(f"{digest}  {path.relative_to(FINAL)}")
    destination.write_text("\n".join(entries) + "\n", encoding="utf-8")


def main() -> int:
    workbook_paths = sorted((FINAL / "Excel").glob("*.xlsx"))
    pdf_paths = sorted((FINAL / "PDF").glob("*.pdf"))
    csv_paths = sorted((FINAL / "Sample Data").rglob("*.csv"))
    screenshot_paths = sorted((FINAL / "Screenshots").rglob("*.png"))
    errors: list[str] = []
    workbooks: list[dict] = []

    for path in workbook_paths:
        result, workbook_errors = audit_workbook(path)
        workbooks.append(result)
        errors.extend(workbook_errors)

    expected_counts = {"workbooks": 4, "pdfs": 4, "sample_csvs": 21, "screenshots": 20}
    actual_counts = {
        "workbooks": len(workbook_paths),
        "pdfs": len(pdf_paths),
        "sample_csvs": len(csv_paths),
        "screenshots": len(screenshot_paths),
    }
    for key, expected in expected_counts.items():
        if actual_counts[key] != expected:
            errors.append(f"Expected {expected} {key}, found {actual_counts[key]}")

    api_key_leaks = find_secret_leaks()
    if api_key_leaks:
        errors.append("Potential API credentials found in final handover")

    result = {
        "status": "PASS" if not errors else "FAIL",
        "expected_counts": expected_counts,
        "actual_counts": actual_counts,
        "workbooks": workbooks,
        "pdfs": [audit_pdf(path) for path in pdf_paths],
        "sample_csvs": [audit_csv(path) for path in csv_paths],
        "api_key_leaks": api_key_leaks,
        "errors": errors,
    }
    destination = FINAL / "Validation" / "release-audit.json"
    destination.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    write_checksums()
    print(json.dumps(result, indent=2))
    return 0 if result["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
